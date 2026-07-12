"""Тесты генерации Excel-экспорта"""

import pytest
from unittest.mock import MagicMock
from datetime import datetime
import openpyxl
from io import BytesIO

from integrations.excel import generate_excel

pytestmark = [pytest.mark.unit]


def make_mock_tx(
    tx_id: int,
    op_type: str = 'EXPENSE',
    amount_kopecks: int = 100000,
    cat_name: str = 'Тест',
    cat_group: str = 'Группа',
    comment: str | None = None,
):
    tx = MagicMock()
    tx.id = tx_id
    tx.type = MagicMock()
    tx.type.value = op_type
    tx.amount_kopecks = amount_kopecks
    tx.comment = comment
    tx.created_at = datetime(2024, 1, 15)
    tx.category = MagicMock()
    tx.category.name = cat_name
    tx.category.group = cat_group
    return tx


class TestGenerateExcel:
    """Тесты структуры Excel-файла"""

    pytestmark = [pytest.mark.unit]

    def test_returns_bytes_and_filename(self):
        txs = [make_mock_tx(1)]
        buf, filename = generate_excel(txs)
        assert isinstance(buf, BytesIO)
        assert filename.endswith('.xlsx')
        assert 'finances_' in filename

    def test_custom_filename_date(self):
        txs = [make_mock_tx(1)]
        _, filename = generate_excel(txs, filename_date='2024-03')
        assert filename == 'finances_2024-03.xlsx'

    def test_excel_has_header_row(self):
        txs = [make_mock_tx(1)]
        buf, _ = generate_excel(txs)
        wb = openpyxl.load_workbook(buf)
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, 6)]
        assert 'Дата' in headers
        assert 'Сумма (руб.)' in headers
        assert 'Категория' in headers

    def test_total_row_present(self):
        """Проверяем, что итоговые строки присутствуют"""
        txs = [
            make_mock_tx(1, op_type='INCOME', amount_kopecks=200000, cat_group='Доходы'),
            make_mock_tx(2, op_type='EXPENSE', amount_kopecks=100000, cat_group='Расходы'),
        ]
        buf, _ = generate_excel(txs)
        wb = openpyxl.load_workbook(buf)
        ws = wb.active

        found_total = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == 'Баланс:':
                    found_total = True
        assert found_total, 'Не найдена строка Баланс:'

    def test_empty_transactions(self):
        """Пустой список — файл всё равно создаётся"""
        buf, filename = generate_excel([])
        assert isinstance(buf, BytesIO)
        wb = openpyxl.load_workbook(buf)
        assert wb.active is not None

    def test_income_and_expense_separation(self):
        """Доходы и расходы попадают в разные строки итогов"""
        income = make_mock_tx(1, op_type='INCOME', amount_kopecks=300000, cat_group='Доходы')
        expense = make_mock_tx(2, op_type='EXPENSE', amount_kopecks=100000, cat_group='Расходы')
        buf, _ = generate_excel([income, expense])
        wb = openpyxl.load_workbook(buf)
        ws = wb.active

        income_row = None
        expense_row = None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == 'Доходы:':
                    income_row = cell.row
                if cell.value == 'Расходы:':
                    expense_row = cell.row

        assert income_row is not None
        assert expense_row is not None
        assert income_row != expense_row

    def test_amount_in_rubles(self):
        """Суммы в файле в рублях, не копейках"""
        txs = [make_mock_tx(1, amount_kopecks=150050)]
        buf, _ = generate_excel(txs)
        wb = openpyxl.load_workbook(buf)
        ws = wb.active

        found = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == 1500.5:
                    found = True
        assert found, 'Сумма 1500.50 не найдена в таблице'
