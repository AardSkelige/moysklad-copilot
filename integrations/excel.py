"""Генерация Excel-отчёта по операциям"""

from io import BytesIO
from collections import defaultdict
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from services.transaction_service import type_label_short


def generate_excel(transactions: list, filename_date: str | None = None) -> tuple[BytesIO, str]:
    """
    Генерирует Excel-отчёт.
    Возвращает (BytesIO с файлом, имя файла).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Операции'

    # Заголовки
    headers = ['Дата', 'Тип', 'Категория', 'Сумма (руб.)', 'Комментарий']
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Данные по группам
    groups = defaultdict(list)
    for tx in transactions:
        group = tx.category.group if tx.category else 'Прочее'
        groups[group].append(tx)

    row = 2
    grand_income = 0
    grand_expense = 0

    for group_name, group_txs in groups.items():
        # Заголовок группы
        group_cell = ws.cell(row=row, column=1, value=group_name)
        group_cell.font = Font(bold=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
        group_cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        row += 1

        group_income = 0
        group_expense = 0

        for tx in group_txs:
            type_label = type_label_short(tx.type)
            amount_rub = tx.amount_kopecks / 100
            cat_name = tx.category.name if tx.category else '?'
            date_str = tx.created_at.strftime('%d.%m.%Y %H:%M')

            ws.cell(row=row, column=1, value=date_str)
            ws.cell(row=row, column=2, value=type_label)
            ws.cell(row=row, column=3, value=cat_name)
            amount_cell = ws.cell(row=row, column=4, value=amount_rub)
            amount_cell.number_format = '#,##0.00'
            ws.cell(row=row, column=5, value=tx.comment or '')

            if tx.type.value == 'INCOME':
                group_income += tx.amount_kopecks
                grand_income += tx.amount_kopecks
            else:
                group_expense += tx.amount_kopecks
                grand_expense += tx.amount_kopecks

            row += 1

        # Итог группы
        ws.cell(row=row, column=3, value='Итого группы:').font = Font(bold=True)
        total_cell = ws.cell(row=row, column=4, value=(group_income - group_expense) / 100)
        total_cell.font = Font(bold=True)
        total_cell.number_format = '#,##0.00'
        row += 1

    # Итоговая строка
    ws.cell(row=row, column=3, value='Доходы:').font = Font(bold=True)
    income_cell = ws.cell(row=row, column=4, value=grand_income / 100)
    income_cell.font = Font(bold=True, color='006400')
    income_cell.number_format = '#,##0.00'
    row += 1

    ws.cell(row=row, column=3, value='Расходы:').font = Font(bold=True)
    expense_cell = ws.cell(row=row, column=4, value=grand_expense / 100)
    expense_cell.font = Font(bold=True, color='8B0000')
    expense_cell.number_format = '#,##0.00'
    row += 1

    ws.cell(row=row, column=3, value='Баланс:').font = Font(bold=True)
    balance_cell = ws.cell(row=row, column=4, value=(grand_income - grand_expense) / 100)
    balance_cell.font = Font(bold=True)
    balance_cell.number_format = '#,##0.00'

    # Ширина колонок
    column_widths = [18, 10, 24, 16, 32]
    for col, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    date_str = filename_date or datetime.now().strftime('%Y-%m-%d')
    filename = f'finances_{date_str}.xlsx'

    return output, filename


def generate_excel_from_ms(
    payments: list[dict],
    category_detail_map: dict,
    filename_date: str | None = None,
) -> tuple[BytesIO, str]:
    """
    Генерирует Excel из списка МС-платежей.
    category_detail_map: {expense_item_href: {'name': str, 'group': str}}
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Операции'

    headers = ['Дата', 'Тип', 'Категория', 'Сумма (руб.)', 'Назначение']
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Группировка: INCOME → 'Доходы', EXPENSE → по группе категории
    groups: dict[str, list[dict]] = defaultdict(list)
    for p in payments:
        if p['op_type'].value == 'INCOME':
            group = 'Доходы'
        elif p['expense_item_href'] and p['expense_item_href'] in category_detail_map:
            group = category_detail_map[p['expense_item_href']]['group'] or 'Прочее'
        else:
            group = 'Прочее'
        groups[group].append(p)

    row = 2
    grand_income = 0
    grand_expense = 0

    for group_name, group_payments in groups.items():
        group_cell = ws.cell(row=row, column=1, value=group_name)
        group_cell.font = Font(bold=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
        group_cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        row += 1

        group_income = 0
        group_expense = 0

        for p in group_payments:
            label = type_label_short(p['op_type'])
            amount_rub = p['amount_kopecks'] / 100
            if p['op_type'].value == 'INCOME':
                cat_name = '—'
            elif p['expense_item_href']:
                cat_name = category_detail_map.get(p['expense_item_href'], {}).get('name', '?')
            else:
                cat_name = '?'
            date_str = p['moment'].strftime('%d.%m.%Y %H:%M')

            ws.cell(row=row, column=1, value=date_str)
            ws.cell(row=row, column=2, value=label)
            ws.cell(row=row, column=3, value=cat_name)
            amount_cell = ws.cell(row=row, column=4, value=amount_rub)
            amount_cell.number_format = '#,##0.00'
            ws.cell(row=row, column=5, value=p['comment'] or '')

            if p['op_type'].value == 'INCOME':
                group_income += p['amount_kopecks']
                grand_income += p['amount_kopecks']
            else:
                group_expense += p['amount_kopecks']
                grand_expense += p['amount_kopecks']

            row += 1

        ws.cell(row=row, column=3, value='Итого группы:').font = Font(bold=True)
        total_cell = ws.cell(row=row, column=4, value=(group_income - group_expense) / 100)
        total_cell.font = Font(bold=True)
        total_cell.number_format = '#,##0.00'
        row += 1

    ws.cell(row=row, column=3, value='Доходы:').font = Font(bold=True)
    income_cell = ws.cell(row=row, column=4, value=grand_income / 100)
    income_cell.font = Font(bold=True, color='006400')
    income_cell.number_format = '#,##0.00'
    row += 1

    ws.cell(row=row, column=3, value='Расходы:').font = Font(bold=True)
    expense_cell = ws.cell(row=row, column=4, value=grand_expense / 100)
    expense_cell.font = Font(bold=True, color='8B0000')
    expense_cell.number_format = '#,##0.00'
    row += 1

    ws.cell(row=row, column=3, value='Баланс:').font = Font(bold=True)
    balance_cell = ws.cell(row=row, column=4, value=(grand_income - grand_expense) / 100)
    balance_cell.font = Font(bold=True)
    balance_cell.number_format = '#,##0.00'

    column_widths = [18, 10, 24, 16, 36]
    for col, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    date_str = filename_date or datetime.now().strftime('%Y-%m-%d')
    filename = f'finances_{date_str}.xlsx'

    return output, filename
